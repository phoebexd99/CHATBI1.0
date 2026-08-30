from __future__ import annotations

import csv
from datetime import date, datetime
from io import BytesIO, StringIO
import json
from pathlib import Path
import re
import sqlite3
from time import perf_counter
from typing import Any, Iterator
from uuid import uuid4

from .db import Database
from .safety import SQLSafetyGate


MAX_UPLOAD_BYTES = 10 * 1024 * 1024
MAX_ROWS_PER_TABLE = 100_000
MAX_COLUMNS = 100
MAX_SHEETS = 10
ALLOWED_COLUMN_ROLES = {"time", "measure", "dimension", "identifier"}


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _quote(identifier: str) -> str:
    return '"' + identifier.replace('"', '""') + '"'


def _literal(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def _format_number(value: Any) -> str:
    if value is None:
        return "—"
    if isinstance(value, float):
        return f"{value:,.2f}"
    if isinstance(value, int):
        return f"{value:,}"
    return str(value)


class DatasetError(ValueError):
    pass


class DatasetService:
    """Local control plane for uploaded spreadsheet datasets.

    The original workbook is never persisted. Normalized values, schema
    metadata, previews, and generated suggestions live in a gitignored SQLite
    file so uploaded user data cannot enter the repository by accident.
    """

    def __init__(self, path: Path):
        self.path = path
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self.database = Database(f"sqlite:///{path}")
        self._initialize()

    def _connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self.path)
        connection.row_factory = sqlite3.Row
        return connection

    def _initialize(self) -> None:
        with self._connect() as connection:
            connection.executescript(
                """
                CREATE TABLE IF NOT EXISTS chatbi_datasets (
                  id TEXT PRIMARY KEY,
                  name TEXT NOT NULL,
                  source_type TEXT NOT NULL,
                  original_filename TEXT NOT NULL,
                  description TEXT NOT NULL,
                  row_count INTEGER NOT NULL,
                  table_count INTEGER NOT NULL,
                  suggestions_json TEXT NOT NULL,
                  created_at TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS chatbi_dataset_tables (
                  id TEXT PRIMARY KEY,
                  dataset_id TEXT NOT NULL,
                  sheet_name TEXT NOT NULL,
                  physical_table TEXT NOT NULL UNIQUE,
                  row_count INTEGER NOT NULL,
                  columns_json TEXT NOT NULL,
                  preview_json TEXT NOT NULL,
                  FOREIGN KEY(dataset_id) REFERENCES chatbi_datasets(id)
                );
                """
            )

    @staticmethod
    def demo_dataset() -> dict[str, Any]:
        return {
            "id": "demo",
            "name": "电商经营演示模板",
            "source_type": "template",
            "status": "ready",
            "description": "内置电商示例，覆盖订单、客户、商品、渠道、营销和库存，用来体验完整智能问数链路。",
            "row_count": 120,
            "table_count": 6,
            "created_at": None,
            "suggestions": [
                "最近 30 天 GMV 是多少？",
                "最近 30 天各渠道退款率",
                "最近 30 天各活动 ROAS 排名",
                "今天各品类可用库存",
            ],
            "tables": [
                {"id": "demo.orders", "sheet_name": "订单", "row_count": 120, "columns": [
                    {"name": "订单日期", "type": "date"}, {"name": "渠道", "type": "text"},
                    {"name": "区域", "type": "text"}, {"name": "成交金额", "type": "real"},
                ], "preview": []},
            ],
        }

    def list_datasets(self) -> list[dict[str, Any]]:
        items = [self.demo_dataset()]
        with self._connect() as connection:
            rows = connection.execute("SELECT * FROM chatbi_datasets ORDER BY created_at DESC").fetchall()
        for row in rows:
            items.append(self._dataset_row(row))
        return items

    def get_dataset(self, dataset_id: str, *, include_internal: bool = False) -> dict[str, Any]:
        if dataset_id == "demo":
            return self.demo_dataset()
        with self._connect() as connection:
            row = connection.execute("SELECT * FROM chatbi_datasets WHERE id = ?", (dataset_id,)).fetchone()
            if not row:
                raise KeyError(dataset_id)
            tables = connection.execute(
                "SELECT * FROM chatbi_dataset_tables WHERE dataset_id = ? ORDER BY id", (dataset_id,)
            ).fetchall()
        result = self._dataset_row(row)
        result["tables"] = [self._table_row(item, include_internal=include_internal) for item in tables]
        return result

    def update_model(self, dataset_id: str, assignments: list[dict[str, str]]) -> dict[str, Any]:
        if dataset_id == "demo":
            raise DatasetError("演示模板使用认证字段模型，不能在上传数据建模页修改")
        with self._connect() as connection:
            dataset = connection.execute("SELECT * FROM chatbi_datasets WHERE id = ?", (dataset_id,)).fetchone()
            if not dataset:
                raise KeyError(dataset_id)
            rows = connection.execute(
                "SELECT * FROM chatbi_dataset_tables WHERE dataset_id = ? ORDER BY id", (dataset_id,)
            ).fetchall()
            table_map = {row["id"]: row for row in rows}
            columns_by_table = {row["id"]: json.loads(row["columns_json"]) for row in rows}
            for assignment in assignments:
                role = assignment["role"]
                if role not in ALLOWED_COLUMN_ROLES:
                    raise DatasetError(f"不支持的字段角色：{role}")
                row = table_map.get(assignment["table_id"])
                if not row:
                    raise DatasetError("字段设置引用了不属于当前数据集的工作表")
                columns = columns_by_table[row["id"]]
                column = next((item for item in columns if item["sql_name"] == assignment["sql_name"]), None)
                if not column:
                    raise DatasetError("字段设置引用了不存在的字段")
                column["role"] = role
            for table_id, columns in columns_by_table.items():
                connection.execute(
                    "UPDATE chatbi_dataset_tables SET columns_json = ? WHERE id = ?",
                    (json.dumps(columns, ensure_ascii=False), table_id),
                )
            refreshed_rows = connection.execute(
                "SELECT * FROM chatbi_dataset_tables WHERE dataset_id = ? ORDER BY id", (dataset_id,)
            ).fetchall()
            suggestion_tables = [{"columns": json.loads(row["columns_json"])} for row in refreshed_rows]
            suggestions = self._suggest_questions(suggestion_tables)
            connection.execute(
                "UPDATE chatbi_datasets SET suggestions_json = ? WHERE id = ?",
                (json.dumps(suggestions, ensure_ascii=False), dataset_id),
            )
            connection.commit()
        return self.get_dataset(dataset_id)

    @staticmethod
    def _dataset_row(row: sqlite3.Row) -> dict[str, Any]:
        return {
            "id": row["id"], "name": row["name"], "source_type": row["source_type"],
            "status": "ready", "description": row["description"],
            "row_count": row["row_count"], "table_count": row["table_count"],
            "created_at": row["created_at"], "suggestions": json.loads(row["suggestions_json"]),
        }

    @staticmethod
    def _table_row(row: sqlite3.Row, *, include_internal: bool = False) -> dict[str, Any]:
        result = {
            "id": row["id"], "sheet_name": row["sheet_name"], "row_count": row["row_count"],
            "columns": json.loads(row["columns_json"]), "preview": json.loads(row["preview_json"]),
        }
        if include_internal:
            result["physical_table"] = row["physical_table"]
        return result

    def upload(self, filename: str, content: bytes, name: str | None = None) -> dict[str, Any]:
        if not content:
            raise DatasetError("文件为空")
        if len(content) > MAX_UPLOAD_BYTES:
            raise DatasetError("文件超过 10 MB 的 MVP 上传限制")
        suffix = Path(filename).suffix.lower()
        if suffix not in {".xlsx", ".csv"}:
            raise DatasetError("当前只支持 .xlsx 和 .csv 文件")
        parsed = self._parse_xlsx(content) if suffix == ".xlsx" else self._parse_csv(content, Path(filename).stem)
        parsed = [item for item in parsed if item[1]]
        if not parsed:
            raise DatasetError("没有找到包含表头和数据行的工作表")

        dataset_id = f"upload_{uuid4().hex[:12]}"
        dataset_name = (name or Path(filename).stem).strip()[:80] or "未命名数据集"
        table_payloads: list[dict[str, Any]] = []
        total_rows = 0
        with self._connect() as connection:
            try:
                for index, (sheet_name, rows) in enumerate(parsed[:MAX_SHEETS], start=1):
                    payload = self._prepare_table(dataset_id, index, sheet_name, rows)
                    table_payloads.append(payload)
                    total_rows += payload["row_count"]
                    definitions = ", ".join(
                        f"{_quote(column['sql_name'])} {column['sqlite_type']}" for column in payload["columns"]
                    )
                    connection.execute(f"CREATE TABLE {_quote(payload['physical_table'])} ({definitions})")
                    placeholders = ", ".join("?" for _ in payload["columns"])
                    connection.executemany(
                        f"INSERT INTO {_quote(payload['physical_table'])} VALUES ({placeholders})",
                        payload["normalized_rows"],
                    )
                suggestions = self._suggest_questions(table_payloads)
                description = (
                    f"从 {filename} 导入，共 {len(table_payloads)} 个工作表、{total_rows:,} 行。"
                    "系统已识别字段类型、样例值和可分析维度。"
                )
                created_at = datetime.now().astimezone().isoformat(timespec="seconds")
                connection.execute(
                    "INSERT INTO chatbi_datasets VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (dataset_id, dataset_name, "excel" if suffix == ".xlsx" else "csv", filename,
                     description, total_rows, len(table_payloads), json.dumps(suggestions, ensure_ascii=False), created_at),
                )
                for payload in table_payloads:
                    public_columns = [{key: value for key, value in item.items() if key != "sqlite_type"} for item in payload["columns"]]
                    connection.execute(
                        "INSERT INTO chatbi_dataset_tables VALUES (?, ?, ?, ?, ?, ?, ?)",
                        (payload["id"], dataset_id, payload["sheet_name"], payload["physical_table"],
                         payload["row_count"], json.dumps(public_columns, ensure_ascii=False),
                         json.dumps(payload["preview"], ensure_ascii=False)),
                    )
                connection.commit()
            except Exception:
                connection.rollback()
                raise
        return self.get_dataset(dataset_id)

    @staticmethod
    def _parse_csv(content: bytes, sheet_name: str) -> list[tuple[str, list[list[Any]]]]:
        decoded = None
        for encoding in ("utf-8-sig", "gb18030"):
            try:
                decoded = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if decoded is None:
            raise DatasetError("CSV 编码无法识别，请使用 UTF-8 或 GB18030")
        return [(sheet_name or "数据", [list(row) for row in csv.reader(StringIO(decoded))])]

    @staticmethod
    def _parse_xlsx(content: bytes) -> list[tuple[str, list[list[Any]]]]:
        try:
            from openpyxl import load_workbook
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except Exception as error:
            raise DatasetError("Excel 文件无法解析，请确认文件未加密且格式为 .xlsx") from error
        result: list[tuple[str, list[list[Any]]]] = []
        try:
            for worksheet in workbook.worksheets[:MAX_SHEETS]:
                rows: list[list[Any]] = []
                for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
                    if row_index > MAX_ROWS_PER_TABLE:
                        raise DatasetError(f"工作表“{worksheet.title}”超过 {MAX_ROWS_PER_TABLE:,} 行限制")
                    rows.append(list(row))
                result.append((worksheet.title, rows))
        finally:
            workbook.close()
        return result

    @staticmethod
    def _prepare_table(dataset_id: str, index: int, sheet_name: str, rows: list[list[Any]]) -> dict[str, Any]:
        header_index = next((i for i, row in enumerate(rows) if any(value not in (None, "") for value in row)), None)
        if header_index is None:
            raise DatasetError(f"工作表“{sheet_name}”为空")
        raw_headers = rows[header_index]
        width = max((len(row) for row in rows[header_index:]), default=len(raw_headers))
        if width > MAX_COLUMNS:
            raise DatasetError(f"工作表“{sheet_name}”超过 {MAX_COLUMNS} 列限制")
        headers: list[str] = []
        used: dict[str, int] = {}
        for column_index in range(width):
            base = str(raw_headers[column_index]).strip() if column_index < len(raw_headers) and raw_headers[column_index] not in (None, "") else f"列{column_index + 1}"
            used[base] = used.get(base, 0) + 1
            headers.append(base if used[base] == 1 else f"{base}_{used[base]}")
        data_rows = [list(row) + [None] * (width - len(row)) for row in rows[header_index + 1:]]
        data_rows = [row[:width] for row in data_rows if any(value not in (None, "") for value in row)]
        if len(data_rows) > MAX_ROWS_PER_TABLE:
            raise DatasetError(f"工作表“{sheet_name}”超过 {MAX_ROWS_PER_TABLE:,} 行限制")

        columns: list[dict[str, Any]] = []
        normalized_columns: list[list[Any]] = []
        for column_index, header in enumerate(headers):
            values = [row[column_index] for row in data_rows]
            inferred = DatasetService._infer_type(values)
            normalized = [DatasetService._normalize_cell(value, inferred) for value in values]
            normalized_columns.append(normalized)
            non_null = [value for value in normalized if value is not None]
            unique_values = list(dict.fromkeys(str(value) for value in non_null))
            columns.append({
                "name": header, "sql_name": f"col_{column_index + 1}", "type": inferred,
                "sqlite_type": "INTEGER" if inferred == "integer" else "REAL" if inferred == "real" else "TEXT",
                "nullable": len(non_null) != len(normalized), "unique_count": len(unique_values),
                "non_null_ratio": round(len(non_null) / len(normalized), 4) if normalized else 0,
                "sample_values": unique_values[:20],
            })
            columns[-1]["role"] = DatasetService._infer_role(
                header, inferred, len(unique_values), len(normalized)
            )
        normalized_rows = [list(values) for values in zip(*normalized_columns)] if normalized_columns else []
        preview = [
            {headers[column_index]: _json_value(row[column_index]) for column_index in range(width)}
            for row in data_rows[:8]
        ]
        return {
            "id": f"{dataset_id}.table_{index}", "sheet_name": sheet_name or f"工作表 {index}",
            "physical_table": f"{dataset_id}_table_{index}", "row_count": len(data_rows),
            "columns": columns, "preview": preview, "normalized_rows": normalized_rows,
        }

    @staticmethod
    def _infer_type(values: list[Any]) -> str:
        non_empty = [value for value in values if value not in (None, "")]
        if not non_empty:
            return "text"
        if all(isinstance(value, (datetime, date)) for value in non_empty):
            return "date"
        if all(isinstance(value, bool) or (isinstance(value, int) and not isinstance(value, bool)) for value in non_empty):
            return "integer"
        if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in non_empty):
            return "real"
        if all(isinstance(value, str) and re.fullmatch(r"[-+]?\d+", value.strip()) for value in non_empty):
            return "integer"
        if all(isinstance(value, str) and re.fullmatch(r"[-+]?(?:\d+\.\d+|\d+)", value.strip()) for value in non_empty):
            return "real"
        date_like = 0
        for value in non_empty[:200]:
            if isinstance(value, str) and re.fullmatch(r"20\d{2}[-/]\d{1,2}[-/]\d{1,2}(?:\s+.*)?", value.strip()):
                date_like += 1
        if date_like == min(len(non_empty), 200):
            return "date"
        return "text"

    @staticmethod
    def _normalize_cell(value: Any, inferred: str) -> Any:
        if value in (None, ""):
            return None
        if inferred == "date":
            if isinstance(value, (datetime, date)):
                return value.isoformat()
            return str(value).strip().replace("/", "-")
        if inferred == "integer":
            return int(value)
        if inferred == "real":
            return float(value)
        return str(value).strip()

    @staticmethod
    def _infer_role(name: str, inferred: str, unique_count: int, row_count: int) -> str:
        normalized_name = name.lower().replace("_", "").replace(" ", "")
        identifier_markers = ("id", "编号", "编码", "单号", "订单号", "客户号", "商品号", "sku")
        looks_like_identifier = any(marker in normalized_name for marker in identifier_markers)
        if looks_like_identifier and unique_count >= max(1, int(row_count * 0.7)):
            return "identifier"
        if inferred == "date":
            return "time"
        if inferred in {"integer", "real"}:
            return "measure"
        return "dimension"

    @staticmethod
    def _suggest_questions(tables: list[dict[str, Any]]) -> list[str]:
        suggestions: list[str] = []
        for table in tables:
            columns = table["columns"]
            numeric = [item for item in columns if item.get("role") == "measure"]
            dates = [item for item in columns if item.get("role") == "time"]
            categories = [item for item in columns if item.get("role") == "dimension" and 1 < item["unique_count"] <= 100]
            if numeric:
                suggestions.extend([f"{numeric[0]['name']}总计是多少？", f"{numeric[0]['name']}平均值是多少？"])
            if numeric and categories:
                suggestions.append(f"按{categories[0]['name']}统计{numeric[0]['name']}合计")
            if numeric and dates:
                suggestions.append(f"按月查看{numeric[0]['name']}趋势")
            if categories:
                suggestions.append(f"各{categories[0]['name']}有多少条记录？")
        return list(dict.fromkeys(suggestions))[:8] or ["这份数据有多少条记录？"]


class UploadedDatasetAnalyzer:
    def __init__(self, service: DatasetService):
        self.service = service
        self.safety = SQLSafetyGate()

    def run(self, dataset_id: str, question: str) -> dict[str, Any]:
        final: dict[str, Any] | None = None
        for event in self.stream(dataset_id, question):
            if event["type"] == "result":
                final = event["state"]
        if final is None:
            raise DatasetError("上传数据分析未返回结果")
        return final

    def stream(self, dataset_id: str, question: str) -> Iterator[dict[str, Any]]:
        trace: list[dict[str, Any]] = []

        def stage(node: str, detail: str, started: float) -> dict[str, Any]:
            item = {"node": node, "status": "ok", "duration_ms": round((perf_counter() - started) * 1000, 2), "detail": detail}
            trace.append(item)
            return {"type": "trace", "trace": item, "node": node}

        started = perf_counter()
        dataset = self.service.get_dataset(dataset_id, include_internal=True)
        if dataset_id == "demo":
            raise DatasetError("演示数据应使用认证指标工作流")
        table = self._select_table(dataset, question)
        yield stage("classify", f"已选择数据集“{dataset['name']}”", started)

        started = perf_counter()
        plan = self._plan(table, question)
        yield stage("extract_entities", f"识别分析字段：{plan['metric_label']}", started)

        started = perf_counter()
        context = self._context(dataset, table, plan)
        yield stage("retrieve", "已读取工作表字段画像和样例值", started)

        started = perf_counter()
        sql = self._sql(table, plan)
        yield stage("generate_sql", "已生成受控聚合查询", started)

        started = perf_counter()
        sql = self.safety.validate(sql, "sqlite", allowed_tables={table["physical_table"]}, allowed_schemas=set())
        self.service.database.explain(sql)
        yield stage("safety", "查询只访问当前上传数据集并已通过只读校验", started)

        started = perf_counter()
        columns, rows = self.service.database.query(sql, max_rows=100)
        display_columns = [plan["group_label"], plan["metric_label"]] if plan["group_sql"] else [plan["metric_label"]]
        yield stage("execute", f"查询完成，返回 {len(rows)} 行", started)

        started = perf_counter()
        chart_type = "line" if plan["group_kind"] == "date" else "bar" if plan["group_sql"] else "metric"
        yield stage("chart", f"已选择{ '趋势图' if chart_type == 'line' else '对比图' if chart_type == 'bar' else '指标卡' }", started)

        started = perf_counter()
        answer = self._answer(plan, rows)
        state = {
            "question": question, "intent": "metric_query", "answer": answer, "sql": sql,
            "columns": display_columns, "rows": rows,
            "chart_spec": {"type": chart_type, "x": display_columns[0] if len(display_columns) > 1 else None, "y": display_columns[-1]},
            "entities": {"metric": plan["metric_label"], "dimensions": [plan["group_label"]] if plan["group_sql"] else [],
                         "time_range": next((item for item in plan["filter_labels"] if "年" in item), "all_time"),
                         "filters": {"applied": plan["filter_labels"]}},
            "retrieval_summary": {"hits": len(context), "top_score": 1.0, "certified_hits": 0, "target_hit": True},
            "evidence": context,
            "insight": {"title": "数据概览", "summary": answer, "highlights": [f"数据集：{dataset['name']}", f"工作表：{table['sheet_name']}", f"返回 {len(rows)} 行结果"]},
            "trace": trace,
        }
        insight_event = stage("insight", "已生成数据结论", started)
        state["trace"] = trace
        yield insight_event
        yield {"type": "result", "state": state}

    @staticmethod
    def _select_table(dataset: dict[str, Any], question: str) -> dict[str, Any]:
        def score(table: dict[str, Any]) -> tuple[int, int]:
            matches = sum(column["name"] in question for column in table["columns"])
            return matches + (2 if table["sheet_name"] in question else 0), table["row_count"]
        return max(dataset["tables"], key=score)

    @staticmethod
    def _plan(table: dict[str, Any], question: str) -> dict[str, Any]:
        numeric = [item for item in table["columns"] if item.get("role", "measure" if item["type"] in {"integer", "real"} else "") == "measure"]
        dates = [item for item in table["columns"] if item.get("role", "time" if item["type"] == "date" else "") == "time"]
        categories = [item for item in table["columns"] if item.get("role", "dimension" if item["type"] == "text" else "") == "dimension" and item["unique_count"] <= 100]
        matched_numeric = next((item for item in sorted(numeric, key=lambda value: -len(value["name"])) if item["name"] in question), None)
        matched_groups = [item for item in [*dates, *categories] if item["name"] in question and item is not matched_numeric]
        asks_count = any(marker in question for marker in ("多少条", "记录数", "行数", "数量")) and not matched_numeric
        operation = "count" if asks_count or not numeric else "avg" if any(marker in question for marker in ("平均", "均值")) else "max" if any(marker in question for marker in ("最大", "最高")) else "min" if any(marker in question for marker in ("最小", "最低")) else "sum"
        metric = matched_numeric or (numeric[0] if numeric else None)
        wants_group = any(marker in question for marker in ("按", "各", "不同", "每", "趋势", "分布", "排名"))
        group = matched_groups[0] if matched_groups and wants_group else dates[0] if dates and "趋势" in question else None
        group_sql = _quote(group["sql_name"]) if group else ""
        group_kind = group["type"] if group else ""
        group_label = group["name"] if group else ""
        if group and group["type"] == "date" and any(marker in question for marker in ("按月", "每月", "月度")):
            group_sql = f"substr({_quote(group['sql_name'])}, 1, 7)"
            group_label = f"{group['name']}（月）"
        metric_sql = "COUNT(*)" if operation == "count" else f"{operation.upper()}({_quote(metric['sql_name'])})"
        operation_label = {"count": "记录数", "sum": "合计", "avg": "平均值", "max": "最大值", "min": "最小值"}[operation]
        metric_label = "记录数" if operation == "count" else f"{metric['name']}{operation_label}"
        top_match = re.search(r"(?:前|top\s*)(\d+)", question, re.IGNORECASE)
        where_clauses: list[str] = []
        filter_labels: list[str] = []
        for column in categories:
            matched_value = next(
                (value for value in sorted(column.get("sample_values", []), key=len, reverse=True) if value and value in question),
                None,
            )
            if matched_value:
                where_clauses.append(f"{_quote(column['sql_name'])} = {_literal(matched_value)}")
                filter_labels.append(f"{column['name']}={matched_value}")
        month_match = re.search(r"(20\d{2})\s*年\s*(1[0-2]|0?[1-9])\s*月", question)
        year_match = re.search(r"(20\d{2})\s*年", question)
        if dates and month_match:
            year, month = int(month_match.group(1)), int(month_match.group(2))
            next_year, next_month = (year + 1, 1) if month == 12 else (year, month + 1)
            date_sql = _quote(dates[0]["sql_name"])
            where_clauses.extend([
                f"date({date_sql}) >= date('{year:04d}-{month:02d}-01')",
                f"date({date_sql}) < date('{next_year:04d}-{next_month:02d}-01')",
            ])
            filter_labels.append(f"{year}年{month}月")
        elif dates and year_match:
            year = int(year_match.group(1))
            date_sql = _quote(dates[0]["sql_name"])
            where_clauses.extend([
                f"date({date_sql}) >= date('{year:04d}-01-01')",
                f"date({date_sql}) < date('{year + 1:04d}-01-01')",
            ])
            filter_labels.append(f"{year}年")
        return {"operation": operation, "metric": metric, "metric_sql": metric_sql, "metric_label": metric_label,
                "group": group, "group_sql": group_sql, "group_kind": group_kind, "group_label": group_label,
                "limit": min(int(top_match.group(1)), 50) if top_match else 20,
                "where_clauses": where_clauses, "filter_labels": filter_labels}

    @staticmethod
    def _sql(table: dict[str, Any], plan: dict[str, Any]) -> str:
        select = f"{plan['metric_sql']} AS metric_value"
        where = f" WHERE {' AND '.join(plan['where_clauses'])}" if plan["where_clauses"] else ""
        group = ""
        order = ""
        if plan["group_sql"]:
            select = f"{plan['group_sql']} AS group_value, {select}"
            group = f" GROUP BY {plan['group_sql']}"
            order = " ORDER BY metric_value DESC"
        return f"SELECT {select} FROM {_quote(table['physical_table'])}{where}{group}{order} LIMIT {plan['limit']}"

    @staticmethod
    def _context(dataset: dict[str, Any], table: dict[str, Any], plan: dict[str, Any]) -> list[dict[str, Any]]:
        names = "、".join(column["name"] for column in table["columns"][:12])
        return [{
            "id": table["id"], "type": "dataset_schema", "title": f"{dataset['name']} · {table['sheet_name']}",
            "text": f"工作表共 {table['row_count']:,} 行，字段包括：{names}。",
            "score": 1.0, "keyword_score": 1.0, "vector_score": 1.0,
            "match_reason": f"问题使用了当前数据集中的“{plan['metric_label']}”字段画像"
                            + (f"，并应用筛选：{'、'.join(plan['filter_labels'])}" if plan["filter_labels"] else ""),
        }]

    @staticmethod
    def _answer(plan: dict[str, Any], rows: list[list[Any]]) -> str:
        if not rows:
            return "当前数据范围内没有可用于该分析的记录。"
        scope = f"在{'、'.join(plan['filter_labels'])}范围内，" if plan["filter_labels"] else ""
        if plan["group_sql"]:
            first = rows[0]
            return f"{scope}已按{plan['group_label']}统计{plan['metric_label']}，共返回 {len(rows)} 组；最高项为“{first[0]}”，结果为 {_format_number(first[1])}。"
        return f"{scope}{plan['metric_label']}为 {_format_number(rows[0][0])}。"
